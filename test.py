from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import load_workbook
from datetime import datetime
import time

# day name
today = datetime.now().strftime("%A") 

# Loading the Excel file
file_path = "C:/Users/ASUS/Desktop/QA_ASSESMENT/Excel.xlsx" 
workbook = load_workbook(file_path)
if today not in workbook.sheetnames:
    print(f"No sheet for {today}. Exiting.")
    exit()
sheet = workbook[today]  

# Selenium WebDriver
driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()))
website = 'https://www.google.com'
driver.get(website)

# Reading keywords
for row_index, row in enumerate(sheet.iter_rows(min_row=3, max_row=12, min_col=3, max_col=3, values_only=True), start=3):
    keyword = row[0].strip() if row[0] else None   
    if not keyword:
        continue  

    # Enter the keyword into Google's search bar
    search_box = driver.find_element(By.NAME, "q")
    search_box.clear()
    search_box.send_keys(keyword)
    time.sleep(2)  

    # Capturing search suggestions
    suggestions = driver.find_elements(By.CSS_SELECTOR, ".erkvQe li span")
    suggestions_text = [s.text for s in suggestions if s.text] 

    longest = max(suggestions_text, key=len) if suggestions_text else "No suggestions"
    shortest = min(suggestions_text, key=len) if suggestions_text else "No suggestions"

    sheet.cell(row=row_index, column=4, value=longest)  
    sheet.cell(row=row_index, column=5, value=shortest)  

workbook.save(file_path)

driver.quit()

print(f"Processing for {today} completed. Results saved to {file_path}.")
